import io
from datetime import date, timedelta, datetime, timezone
from fastapi import APIRouter, Depends, HTTPException, Response
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from sqlalchemy import select

from database import get_db
from models import Item, Batch, WithdrawalLog
from auth import require_manager

router = APIRouter(prefix="/api/reports", tags=["reports"])


async def _load_items(db: AsyncSession):
    result = await db.execute(
        select(Item).options(
            selectinload(Item.category), selectinload(Item.room),
            selectinload(Item.batches).selectinload(Batch.withdrawals),
        )
    )
    return result.scalars().all()


async def _load_withdrawals(db: AsyncSession, date_from: date, date_to: date):
    result = await db.execute(
        select(WithdrawalLog)
        .options(selectinload(WithdrawalLog.user), selectinload(WithdrawalLog.batch).selectinload(Batch.item))
        .where(WithdrawalLog.withdrawn_at >= date_from, WithdrawalLog.withdrawn_at <= date_to + timedelta(days=1))
        .order_by(WithdrawalLog.withdrawn_at.desc())
    )
    return result.scalars().all()


async def _load_active_batches_with_expiry(db: AsyncSession):
    result = await db.execute(
        select(Batch)
        .options(selectinload(Batch.item), selectinload(Batch.withdrawals))
        .where(Batch.is_active == True, Batch.expiry_date.isnot(None))  # noqa: E712
        .order_by(Batch.expiry_date)
    )
    return result.scalars().all()


def _last_priced_batch(item: Item):
    priced = [b for b in item.batches if float(b.unit_price) > 0]
    return max(priced, key=lambda b: b.purchase_date) if priced else None


def _generate_excel(items, withdrawals, batches_with_expiry):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    NAVY = "1a3c5e"; WHITE = "FFFFFF"; LIGHT = "f0f4f8"
    HDR_FILL = PatternFill("solid", fgColor=NAVY)
    HDR_FONT = Font(color=WHITE, bold=True, size=11)
    THIN = Side(style='thin', color='cccccc')
    BRD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def hdr(ws, cols):
        for i, h in enumerate(cols, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.fill = HDR_FILL; c.font = HDR_FONT
            c.alignment = Alignment(horizontal='center', vertical='center'); c.border = BRD
        ws.row_dimensions[1].height = 22

    def autowidth(ws, minw=12):
        for col in ws.columns:
            w = max((len(str(c.value or '')) for c in col), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max(w + 4, minw)

    def stripe(ws, start=2):
        alt = PatternFill("solid", fgColor=LIGHT)
        for i, row in enumerate(ws.iter_rows(min_row=start), start=start):
            fill = alt if i % 2 == 0 else PatternFill()
            for cell in row:
                cell.fill = fill; cell.border = BRD
                cell.alignment = Alignment(vertical='center')

    SM = {'good': 'Well Stocked', 'moderate': 'Moderate', 'low': 'Low Stock', 'out': 'Out of Stock'}
    EM = {'none': '—', 'ok': 'OK', 'warning': 'Expiring Soon', 'critical': 'Expiring This Week', 'expired': 'EXPIRED'}

    ws1 = wb.active; ws1.title = "Inventory Snapshot"
    hdr(ws1, ['Item', 'Brand', 'Category', 'Room', 'Unit', 'Total Stock', 'Reorder Threshold', 'Status', 'Earliest Expiry', 'Expiry Status'])
    for item in items:
        ws1.append([
            item.name, item.brand or '—', item.category.name if item.category else '—',
            item.room.name if item.room else '—', item.unit_type, item.total_quantity(),
            float(item.reorder_threshold), SM.get(item.stock_status(), '—'),
            str(item.earliest_expiry() or '—'), EM.get(item.expiry_status(), '—'),
        ])
    stripe(ws1); autowidth(ws1)

    ws2 = wb.create_sheet("Withdrawal History")
    hdr(ws2, ['Date & Time', 'Item', 'Brand', 'Batch #', 'Qty Taken', 'Remaining After', 'Taken By', 'Purpose', 'Notes'])
    for log in withdrawals:
        ws2.append([
            log.withdrawn_at.strftime('%Y-%m-%d %H:%M'), log.batch.item.name, log.batch.item.brand or '—',
            f"#{log.batch.id}", float(log.quantity_taken), float(log.quantity_remaining_after),
            log.user.full_name() if log.user else 'Unknown', log.purpose or '—', log.notes or '—',
        ])
    if ws2.max_row == 1:
        ws2.append(['No withdrawals in this period'] + [''] * 8)
    stripe(ws2); autowidth(ws2)

    ws3 = wb.create_sheet("Expiry Report")
    hdr(ws3, ['Item', 'Brand', 'Batch #', 'Purchase Date', 'Expiry Date', 'Days Until Expiry', 'Remaining Units', 'Status'])
    today = date.today()
    RED_FILL = PatternFill("solid", fgColor="ffe0e0")
    AMBER_FILL = PatternFill("solid", fgColor="fff3cd")
    for batch in batches_with_expiry:
        days = (batch.expiry_date - today).days
        status = 'EXPIRED' if days < 0 else ('Critical (<7d)' if days <= 7 else ('Soon (<30d)' if days <= 30 else 'OK'))
        ws3.append([batch.item.name, batch.item.brand or '—', f"#{batch.id}",
            str(batch.purchase_date), str(batch.expiry_date), days, batch.remaining_units(), status])
        row = ws3.max_row
        fill = RED_FILL if days < 0 else (AMBER_FILL if days <= 7 else None)
        if fill:
            for col in range(1, 9):
                ws3.cell(row=row, column=col).fill = fill
    if ws3.max_row == 1:
        ws3.append(['No batches with expiry dates'] + [''] * 7)
    stripe(ws3); autowidth(ws3)

    ws4 = wb.create_sheet("Restock Estimate")
    hdr(ws4, ['Item', 'Brand', 'Category', 'Current Stock', 'Reorder Threshold', 'Unit', 'Last Price (₦)', 'Est. Restock Cost (₦)', 'Status'])
    total_est = 0
    for item in [i for i in items if i.stock_status() in ('low', 'out')]:
        lb = _last_priced_batch(item)
        price = float(lb.unit_price) if lb else 0
        total_est += price
        ws4.append([item.name, item.brand or '—', item.category.name if item.category else '—',
            item.total_quantity(), float(item.reorder_threshold), item.unit_type, price, price,
            'Out of Stock' if item.stock_status() == 'out' else 'Low Stock'])
    if ws4.max_row == 1:
        ws4.append(['All items adequately stocked'] + [''] * 8)
    stripe(ws4)
    tr = ws4.max_row + 1
    ws4.cell(row=tr, column=7, value='TOTAL ESTIMATE').font = Font(bold=True)
    ws4.cell(row=tr, column=8, value=total_est).font = Font(bold=True)
    autowidth(ws4)

    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out


def _generate_pdf(items, withdrawals, batches_with_expiry, date_from, date_to):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, HRFlowable
    from reportlab.lib.enums import TA_CENTER

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=1.5*cm, bottomMargin=1.5*cm, leftMargin=1.5*cm, rightMargin=1.5*cm)
    NAVY = colors.HexColor('#1a3c5e')
    LIGHT = colors.HexColor('#f0f4f8')
    RED = colors.HexColor('#dc2626')
    AMBER = colors.HexColor('#d97706')
    GREEN = colors.HexColor('#16a34a')

    styles = getSampleStyleSheet()
    TITLE = ParagraphStyle('T', fontSize=20, textColor=NAVY, spaceAfter=4, alignment=TA_CENTER, fontName='Helvetica-Bold')
    SUB = ParagraphStyle('S', fontSize=11, textColor=colors.grey, spaceAfter=10, alignment=TA_CENTER)
    SEC = ParagraphStyle('SC', fontSize=14, textColor=NAVY, spaceBefore=12, spaceAfter=6, fontName='Helvetica-Bold')

    def tbl_style(extra=None):
        base = [
            ('BACKGROUND', (0, 0), (-1, 0), NAVY), ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, LIGHT]), ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#cccccc')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'), ('TOPPADDING', (0, 0), (-1, -1), 4), ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]
        if extra: base.extend(extra)
        return TableStyle(base)

    story = [Spacer(1, 0.8*cm)]
    story.append(Paragraph("Home Inventory", TITLE))
    story.append(Paragraph("Comprehensive Inventory Report", SUB))
    story.append(Paragraph(f"Period: {date_from.strftime('%d %b %Y')} — {date_to.strftime('%d %b %Y')}", SUB))
    story.append(Paragraph(f"Generated: {datetime.now(timezone.utc).strftime('%d %b %Y, %I:%M %p')}", SUB))
    story.append(HRFlowable(width="100%", thickness=2, color=NAVY))
    story.append(Spacer(1, 0.4*cm))

    story.append(Paragraph("1. Inventory Snapshot", SEC))
    SM = {'good': 'Good', 'moderate': 'Moderate', 'low': 'Low', 'out': 'OUT'}
    data1 = [['Item', 'Brand', 'Category', 'Room', 'Unit', 'Stock', 'Threshold', 'Status', 'Earliest Expiry']]
    extra1 = []
    for i, item in enumerate(items, 1):
        s = item.stock_status()
        data1.append([item.name[:24], (item.brand or '—')[:14],
            (item.category.name if item.category else '—')[:14], (item.room.name if item.room else '—')[:14],
            item.unit_type[:10], str(item.total_quantity()),
            str(item.reorder_threshold), SM.get(s, '—'), str(item.earliest_expiry() or '—')])
        if s == 'out': extra1.append(('TEXTCOLOR', (7, i), (7, i), RED))
        elif s == 'low': extra1.append(('TEXTCOLOR', (7, i), (7, i), AMBER))
        elif s == 'good': extra1.append(('TEXTCOLOR', (7, i), (7, i), GREEN))
    t1 = Table(data1, repeatRows=1); t1.setStyle(tbl_style(extra1))
    story.append(t1); story.append(PageBreak())

    story.append(Paragraph("2. Withdrawal History", SEC))
    data2 = [['Date', 'Item', 'Batch #', 'Qty Taken', 'Remaining', 'Taken By', 'Purpose']]
    for log in withdrawals:
        data2.append([log.withdrawn_at.strftime('%d/%m/%y %H:%M'), log.batch.item.name[:22],
            f"#{log.batch.id}", str(log.quantity_taken), str(log.quantity_remaining_after),
            (log.user.full_name() if log.user else 'Unknown')[:18], (log.purpose or '—')[:28]])
    if len(data2) == 1: data2.append(['No withdrawals in this period', '', '', '', '', '', ''])
    t2 = Table(data2, repeatRows=1); t2.setStyle(tbl_style())
    story.append(t2); story.append(PageBreak())

    story.append(Paragraph("3. Expiry Report", SEC))
    today = date.today()
    data3 = [['Item', 'Brand', 'Batch #', 'Purchase Date', 'Expiry Date', 'Days Left', 'Remaining Units', 'Status']]
    extra3 = []
    for i, batch in enumerate(batches_with_expiry, 1):
        days = (batch.expiry_date - today).days
        status = 'EXPIRED' if days < 0 else ('Critical' if days <= 7 else ('Soon' if days <= 30 else 'OK'))
        data3.append([batch.item.name[:22], (batch.item.brand or '—')[:12],
            f"#{batch.id}", str(batch.purchase_date), str(batch.expiry_date), str(days), str(batch.remaining_units()), status])
        if days < 0:
            extra3.append(('TEXTCOLOR', (7, i), (7, i), RED))
            extra3.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#ffe0e0')))
        elif days <= 7:
            extra3.append(('TEXTCOLOR', (7, i), (7, i), AMBER))
    if len(data3) == 1: data3.append(['No batches with expiry dates', '', '', '', '', '', '', ''])
    t3 = Table(data3, repeatRows=1); t3.setStyle(tbl_style(extra3))
    story.append(t3); story.append(PageBreak())

    story.append(Paragraph("4. Restock Estimate", SEC))
    low_items = [i for i in items if i.stock_status() in ('low', 'out')]
    data4 = [['Item', 'Brand', 'Category', 'Current Stock', 'Threshold', 'Unit', 'Last Price (₦)', 'Est. Cost (₦)', 'Status']]
    total_est = 0
    for item in low_items:
        lb = _last_priced_batch(item)
        price = float(lb.unit_price) if lb else 0
        total_est += price
        data4.append([item.name[:22], (item.brand or '—')[:12],
            (item.category.name if item.category else '—')[:12], str(item.total_quantity()),
            str(item.reorder_threshold), item.unit_type[:10],
            f"{price:,.0f}", f"{price:,.0f}", 'Out' if item.stock_status() == 'out' else 'Low'])
    if not low_items: data4.append(['All items are adequately stocked.', '', '', '', '', '', '', '', ''])
    last = len(data4)
    data4.append(['', '', '', '', '', '', 'TOTAL ESTIMATE:', f"N{total_est:,.0f}", ''])
    t4 = Table(data4, repeatRows=1)
    t4.setStyle(tbl_style([('FONTNAME', (0, last), (-1, last), 'Helvetica-Bold'), ('TEXTCOLOR', (6, last), (7, last), NAVY)]))
    story.append(t4)

    doc.build(story); buf.seek(0)
    return buf


@router.get("/generate")
async def generate_report(
    format: str = "pdf",
    period: str = "month",
    date_from: str | None = None,
    date_to: str | None = None,
    db: AsyncSession = Depends(get_db),
    _=Depends(require_manager),
):
    today = date.today()
    if period == "week":
        d_from, d_to = today - timedelta(days=7), today
    elif period == "3months":
        m = today.month - 3; y = today.year
        if m <= 0:
            m += 12; y -= 1
        d_from, d_to = today.replace(year=y, month=m, day=1), today
    elif period == "custom" and date_from and date_to:
        try:
            d_from, d_to = date.fromisoformat(date_from), date.fromisoformat(date_to)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format, use YYYY-MM-DD")
    else:
        d_from, d_to = today.replace(day=1), today

    items = await _load_items(db)
    withdrawals = await _load_withdrawals(db, d_from, d_to)
    batches_with_expiry = await _load_active_batches_with_expiry(db)

    if format == "excel":
        out = _generate_excel(items, withdrawals, batches_with_expiry)
        fn = f"home_inventory_report_{d_from}_{d_to}.xlsx"
        return Response(
            content=out.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{fn}"'},
        )
    buf = _generate_pdf(items, withdrawals, batches_with_expiry, d_from, d_to)
    fn = f"home_inventory_report_{d_from}_{d_to}.pdf"
    return Response(
        content=buf.read(),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fn}"'},
    )
